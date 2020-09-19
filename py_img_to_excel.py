from openpyxl.workbook import Workbook#导入Workbook库用与操作Execl工作簿
from openpyxl.styles import PatternFill, Color#导入PatternFill,Color库用与操作Execl单元格
from PIL import Image#导入Image库用与操作图片文件
import datetime

#把一个整数值转换成26进制字符串
#因为execl单元格的行坐标是26进制的, 比如"A", "Z", "AA", "AZ"
def dec_to_base26(d):
    s = ""
    m = 0
    while d > 0:
        m = d % 26
        if m == 0:
            m = 26
        s = "{0:c}{1:s}".format(m+64, s)
        d = (d - m) // 26
    return s

#把一个26进制字符串转换成整数值
def base26_to_dec(s):
    d = 0
    j = 1
    st = s.upper()
    for x in range(0, len(st))[::-1]:
        c = ord(st[x])
        if c < 65 and c > 90:
            return 0
        d += (c - 64) * j
        j *= 26
    return d

#把一个整数坐标转换成Execl坐标
#Execl坐标的行坐标是26进制的, 列坐标是10进制的,比如(AA, 100)
def decxy_to_excelxy(x, y):
    return("{0:s}{1:d}".format(dec_to_base26(x), y))

#把像素点的rgb值转换成Execl支持的十六进制字符串, 形如 "AARRGGBB",
#其中AA表示透明度,这里设置为0, 比如 "00FF55FF"
def pixel_to_xrgbstr(pix):
    return ("00{0:02X}{1:02X}{2:02X}".format(pix[0], pix[1], pix[2]))

#图片转Execl函数, imgName 表示带全路径的图片名
def image_to_excel(imgName):

    #创建一个 excel 工作簿
    wb = Workbook()
    ws = wb.active

    #打开图片文件文件
    print("Open Image File [{0}]".format(imgName))
    try:
        img = Image.open(imgName)
    except:
        print("Error to Open [{0}]!!!".format(imgName))

    #判断图片文件的格式, 这里必须为"RGB"格式, 如果不是"RGB"格式, 
    #则用convert函数转换成"RGB"格式.
    if "RGB" == img.mode:
        print("Size{0},Format({1}),Color({2})".format(img.size, img.format, img.mode))
    else:
        print("Not a RGB image file!!!")
        img = img.convert("RGB")
        print("Convert to RGB Success!!!")

    #获取图片文件宽和高
    width = img.size[0]#宽度
    height = img.size[1]#高度

    zoom = 0

    #如果图片文件大于 400*400像素,则对图片进行缩放,缩放比例依照宽度和高度中的最大值
    PIC_MAXSIZE = 640

    if PIC_MAXSIZE != 0:
        if width >= height:
            maxsize = width
        else:
            maxsize = height

        if maxsize >= PIC_MAXSIZE:
            zoom = maxsize / PIC_MAXSIZE
            width = int(width / zoom)
            height = int(height / zoom)
            img = img.resize((width, height))
            print("Image Size too large, Resize to", img.size)

    index = 0

    print("Start Process!")
    for w in range(width):#遍历图片的宽度,[0, width)
        #显示处理进度
        index += 1
        print("#", end="")
        if index >= 60:#大于60换行
            index = 0
            print("")

        for h in range(height):#遍历图片的高度[0, height)
            pixel = img.getpixel((w, h))#获取图片当前坐标点的像素值
            loc = decxy_to_excelxy(w+1, h+1)#把整数坐标转换成Execl坐标(字符串)
            #print("LOC", loc)
            c = ws[loc]#选中当前图片像素点坐标对应的Execl单元格
            col = pixel_to_xrgbstr(pixel)#把当前图片像素点的颜色转换成Execl单元格的颜色(字符串)
            #print("COL", col)
            #用当前图片像素点的颜色填充单元格底色
            cfill = PatternFill(fill_type="solid", start_color=col)
            c.fill = cfill
            #把单元格的宽设置为1,高设置为6, 这样单元格看上去就是一个(小)正方形
            ws.column_dimensions[dec_to_base26(w+1)].width = 1
            ws.row_dimensions[h+1].height = 6

    print("\nProcess Done!")
    #获取当前时间,转换成字符串
    timenow = datetime.datetime.now()
    timestr = timenow.strftime("%Y-%m-%d-%H-%M-%S")
    #生成的Execl文件用<原图片文件名+ 当前时间字符串+ ".xlsx"后缀>作为文件名
    namestr = "{0}-{1}.xlsx".format(imgName, timestr)
    print("Save File As [{0}]".format(namestr))
    #保存新生成的Execl文件
    wb.save(namestr)
    print("Save Done!")

name = input("Please Input Image File Name:")
print("Start......")
try:
    image_to_excel(name)
except:
    print("Error!!!!!!")
print("Over......")


